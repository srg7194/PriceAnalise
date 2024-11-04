import gspread
import openai
import clickhouse_connect
import clickhouse_driver
import clickhouse_sqlalchemy
import googleapiclient
import google.auth
import google.oauth2
import googleapiclient.errors
import googleapiclient.http
import pandas as pd
import numpy as np
import sqlalchemy
import psycopg2
import pymysql
import sqlite3
import json
import yaml
import os
import io
import requests
import datetime
import importlib
import subprocess
import sys
import logging
import time
import glob
import tabulate
import re
import transliterate
import shutil
import multiprocessing_logging
import multiprocessing
import telegram
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ParseMode, Bot
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler, CallbackContext, MessageHandler, Filters, \
    JobQueue, ConversationHandler
import threading
from threading import Lock
import warnings
import inspect
from colorama import Fore, Style, init
# Импорты Google
import pprint

import gspread
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from google.oauth2 import service_account
from googleapiclient.errors import HttpError
from gspread.exceptions import GSpreadException
from oauth2client.service_account import ServiceAccountCredentials

# Импорты OpenPyXL
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Импорты SQLAlchemy и ClickHouse
from sqlalchemy import Table, Column, Integer, String
from sqlalchemy import create_engine, exc, text, orm, dialects
from sqlalchemy.orm import sessionmaker
import clickhouse_sqlalchemy
from clickhouse_driver import Client
import clickhouse_connect
import psycopg2
import pymysql
import sqlite3

# Импорты YAML, JSON
import json
import yaml
from yaml import SafeLoader

# Импорты Pandas, Numpy
import pandas as pd
import numpy as np

# Остальные импорты
import os
import io
import requests
from datetime import datetime

import datetime
import importlib
import subprocess
import sys
import logging
import time
import glob

from colorama import Fore, Style, init

import importlib
import subprocess
import sys
import inspect
import os

import tabulate
import re
from functools import wraps

import transliterate
import shutil
# from googletrans import Translator
# from mtranslate import translate as mtranslate_translate
from translate import Translator

import pandas as pd
from google.oauth2.credentials import Credentials
from gspread_dataframe import set_with_dataframe
import multiprocessing_logging
import multiprocessing

from openai import OpenAI
import re
import xlsxwriter

from memory_profiler import profile

warnings.filterwarnings('ignore')

""" Логирование """


def log_decorator(logger, good='DEBUG', bad='ERROR', write_stage=False, storage_hours=1, debug=False):
    def _choise(logger, log_level):
        if log_level == 'DEBUG':
            return logger.log_debug
        elif log_level == "INFO":
            return logger.log_info
        elif log_level == "WARNING":
            return logger.log_warning
        elif log_level == "ERROR":
            return logger.log_error
        elif log_level == "CRITICAL":
            return logger.log_critical
        else:
            logger.log_error(f'''error: Incorrect logger choise. <<{log_level}>> type selected''')
            return logger.log_debug

    good_level_logger = _choise(logger, good)
    bad_level_logger = _choise(logger, bad)

    def decorator(func):
        # Обёртка для функции или метода, добавляющая логирование перед и после выполнения
        def wrapper(*args, **kwargs):
            def write_stage_to_db(logger, write_stage, message, parameter, state, timelag, error, args, kwargs):
                # if write_stage:
                #     state = {
                #         'datetime': datetime.datetime.now(),
                #         'name': str(logger.name),
                #         'stage': str(write_stage),
                #         'worker': str(message),
                #         'parameter': str(parameter),
                #         'state': str(state),
                #         'error': str(error),
                #         'timelag': str(timelag),
                #         # 'args': str(args),
                #         # 'kwargs': str(kwargs)
                #
                #     }
                #     _dir = 'files'
                #     if not FileUtils.file_exists(_dir):
                #         FileUtils.create_directory(_dir)
                #     db_name = f'{_dir}/system.db'
                #
                #     try:
                #         # Создание/подключение к БД
                #         connection = sqlite3.connect(db_name)
                #         connection.close()
                #
                #         ''' Период '''
                #         table_name = 'state_history'
                #         connection = sqlite3.connect(db_name)
                #         cursor = connection.cursor()
                #
                #         # Проверка наличия
                #         query = f"""SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'"""
                #         if cursor.execute(query).fetchall():
                #             # Чистка таблицы
                #             min_date = datetime.datetime.now() - datetime.timedelta(hours=storage_hours)
                #             query = f"""DELETE FROM {table_name} WHERE datetime(datetime) < datetime('{min_date}')"""
                #             cursor.execute(query)
                #
                #         connection.commit()
                #         connection.close()
                #
                #         # Запись данных
                #         engine = create_engine(f'sqlite:///{db_name}')
                #         pd.DataFrame([state]).to_sql(table_name, engine, index=False, if_exists='append')
                #     except Exception as err:
                #         bad_level_logger(f'text: Error logging to database. {err}')
                #     # ''' Последнее состояние '''
                #     # table_name = 'state_last'
                #     # connection = sqlite3.connect(db_name)
                #     # cursor = connection.cursor()
                #     #
                #     # # Проверка наличия
                #     # query = f"""SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'"""
                #     # if cursor.execute(query).fetchall():
                #     #     # Чистка таблицы
                #     #     query = f"""DELETE FROM {table_name} WHERE name = '{state['name']}' AND stage = '{state['stage']}' AND worker = '{state['worker']}'"""
                #     #     cursor.execute(query)
                #     # connection.commit()
                #     # connection.close()
                #     # # Запись данных
                #     # engine = create_engine(f'sqlite:///{db_name}')
                #     # pd.DataFrame([state]).to_sql(table_name, engine, index=False, if_exists='append')
                #     #
                #     # ''' Ошибки за 7 дней '''
                #     # table_name = 'error'
                #     # connection = sqlite3.connect(db_name)
                #     # cursor = connection.cursor()
                #     #
                #     # # Проверка наличия
                #     # query = f"""SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'"""
                #     # if cursor.execute(query).fetchall():
                #     #     # Чистка таблицы
                #     #     min_date = datetime.datetime.now() - datetime.timedelta(weeks=1)
                #     #     query = f"""DELETE FROM {table_name} WHERE datetime(datetime) < datetime('{min_date}')"""
                #     #     cursor.execute(query)
                #     # connection.commit()
                #     # connection.close()
                #     # # Запись данных
                #     # if state['state'] == 'error':
                #     #     engine = create_engine(f'sqlite:///{db_name}')
                #     #     pd.DataFrame([state]).to_sql(table_name, engine, index=False, if_exists='append')
                return

            # Проверяем, является ли первый аргумент экземпляром класса
            if args and hasattr(args[0], '__class__') and hasattr(args[0], '__dict__'):
                instance = args[0]
                class_name = instance.__class__.__name__
                method_name = func.__name__
                message = f"Classmethod <<{class_name}.{method_name}>>"
            # Если нет, то определяем класс это или функция
            else:
                message = f"Function <<{func.__name__}>>"
                if 'class' in str(func).lower():
                    message = f"Class <<{func.__name__}>>"
            console_message = f"'worker': '{message}'"

            # Извлечение self из args
            parameter = False
            try:
                self = args[0] if args else None
                found = False
                if self:
                    table_name = getattr(self, 'table_name', None)
                    if table_name:
                        found = True
                        console_message += f"'table_name': {table_name}"
                        parameter = f"'table_name': {table_name}"

                if not found:
                    if message == 'Classmethod <<DataStatusUpdater.__get_info>>':
                        for arg in args:
                            if arg != self and type(arg) == str:
                                for el in ['.json', '.xlsx', '.yaml', '.yaml']:
                                    if el in arg:
                                        console_message += f", 'file': {arg}"
                                        parameter = f"'file': {arg}"
                                        break
            except:
                pass

            if not parameter:
                parameter = ''

            good_level_logger(f'{{stage: {write_stage}, {console_message}, status: start}}')
            write_stage_to_db(logger, write_stage, message, parameter, state='start', timelag='', error='',
                              args=args, kwargs=kwargs)
            st = datetime.datetime.now()

            delay = 0.01

            # Выполняем функцию и возвращаем результат
            if debug:
                result = func(*args, **kwargs)
                good_level_logger(f'stage: {write_stage}, '
                                  f'{console_message}, '
                                  f'status: done, '
                                  f'timelag: {datetime.datetime.now() - st}')
                write_stage_to_db(logger, write_stage, message, parameter, state='done',
                                  timelag=datetime.datetime.now() - st,
                                  error='', args=args, kwargs=kwargs)
            else:
                try:
                    result = func(*args, **kwargs)
                    time.sleep(delay)
                    good_level_logger(f'stage: {write_stage}, '
                                      f'{console_message}, '
                                      f'status: done, '
                                      f'timelag: {datetime.datetime.now() - st}')
                    write_stage_to_db(logger, write_stage, message, parameter, state='done',
                                      timelag=datetime.datetime.now() - st, error='', args=args, kwargs=kwargs)

                except Exception as err:
                    result = f'Error! {err}'
                    time.sleep(delay)
                    bad_level_logger(f'stage: {write_stage}, {console_message}, '
                                     f'status: error, timelag: {datetime.datetime.now() - st}, '
                                     f'text: {err}')
                    write_stage_to_db(logger, write_stage, message, parameter, state='error',
                                      timelag=datetime.datetime.now() - st, error=err, args=args, kwargs=kwargs)
            return result

        return wrapper

    return decorator


class CustomLogger:
    """
    Класс CustomLogger предоставляет настраиваемый логгер с поддержкой цветного вывода в консоль.

    Аргументы:
    - name: имя логгера
    - log_file_path: путь к директории, где будет сохранен файл лога
    - console_log: опция вывода в консоль (по умолчанию True)
    - file_log: опция сохранения лога в файл (по умолчанию False)

    Уровни логирования
    - Debug (10): самый низкий уровень логирования, предназначенный для отладочных сообщений, для вывода диагностической информации о приложении.
    - Info (20): этот уровень предназначен для вывода данных о фрагментах кода, работающих так, как ожидается.
    - Warning (30): этот уровень логирования предусматривает вывод предупреждений, он применяется для записи сведений о событиях, на которые программист обычно обращает внимание. Такие события вполне могут привести к проблемам при работе приложения. Если явно не задать уровень логирования — по умолчанию используется именно warning.
    - Error (40): этот уровень логирования предусматривает вывод сведений об ошибках — о том, что часть приложения работает не так как ожидается, о том, что программа не смогла правильно выполниться.
    - Critical (50): этот уровень используется для вывода сведений об очень серьёзных ошибках, наличие которых угрожает нормальному функционированию всего приложения. Если не исправить такую ошибку — это может привести к тому, что приложение прекратит работу.
    """

    def __init__(self, name, level=logging.DEBUG, log_file_path=False, console_log=True, file_log=False):
        """
        Инициализирует экземпляр класса CustomLogger.

        Пример использования:
        logger = CustomLogger(name='example_logger', log_file_path='/path/to/logs', console_log=True, file_log=True)
        logger.log_info('This is an informational message.')

        Аргументы:
        - name: имя логгера
        - log_file_path: путь к директории, где будет сохранен файл лога
        - console_log: опция вывода в консоль (по умолчанию True)
        - file_log: опция сохранения лога в файл (по умолчанию False)
        """
        # Инициализация логгера с настройками
        self.logger = logging.getLogger(name)
        self.logger.setLevel(level)

        self.console_log = console_log
        self.file_log = file_log
        self.name = name

        # Генерация пути к файлу лога с учетом времени
        log_file_path = f'{log_file_path}/log_{datetime.datetime.now().replace(microsecond=0)}.log'

        # Форматтер
        formatter = logging.Formatter(f"{{'time': '%(asctime)s', 'logger': '%(name)s', 'level': '%(levelname)s', "
                                      f"'file': '%(module)s', 'string': %(lineno)d, 'message': '%(message)s'}}")

        # Обработчик для консоли
        if self.console_log:
            console_handler = logging.StreamHandler(sys.stdout)
            console_handler.setFormatter(formatter)
            self.logger.addHandler(console_handler)

        # Обработчик для файла
        if self.file_log:
            file_handler = logging.FileHandler(log_file_path)
            file_handler.setFormatter(formatter)
            self.logger.addHandler(file_handler)

        # Инициализация colorama для цветного вывода в консоль
        init()

    def set_log_level(self, level):
        # Установка уровня логирования для логгера
        self.logger.setLevel(level)

    @staticmethod
    def add_info_to_msg(msg, add=False):
        if add:
            # Добавление дополнительной информации к сообщению лога
            add = {f'{k}': f'{v}' for k, v in add.items()}
            add = str(add).replace('{', '').replace('}', '') if add else ''  # .replace("'", '')
            if not msg or msg in (''):
                msg = False
            msg = f'{msg}, {add}'
        if msg[:1] != '{':
            return str({'text': msg})  # .replace("'", '')
        return msg

    def log_debug(self, message, add=False):
        # Логирование сообщения с уровнем DEBUG
        self.logger.debug(f"{Fore.CYAN}{Style.BRIGHT}{self.add_info_to_msg(message, add)}{Style.RESET_ALL}")

    def log_info(self, message, add=False):
        # Логирование сообщения с уровнем INFO
        self.logger.info(f"{Fore.GREEN}{self.add_info_to_msg(message, add)}{Style.RESET_ALL}")

    def log_warning(self, message, add=False):
        # Логирование сообщения с уровнем WARNING
        self.logger.warning(f"{Fore.YELLOW}{self.add_info_to_msg(message, add)}{Style.RESET_ALL}")

    def log_error(self, message, add=False):
        # Логирование сообщения с уровнем ERROR
        self.logger.error(f"{Fore.RED}{Style.BRIGHT}{self.add_info_to_msg(message, add)}{Style.RESET_ALL}")

    def log_critical(self, message, add=False):
        # Логирование сообщения с уровнем CRITICAL
        self.logger.critical(f"{Fore.MAGENTA}{Style.BRIGHT}{self.add_info_to_msg(message, add)}{Style.RESET_ALL}")


level_logging = logging.INFO
logger = CustomLogger(name=os.path.basename('AdditionalMethods'), level=level_logging)

""" Вспомогательные функции """


@log_decorator(logger=logger)
def df_print(df):
    """
    Выводит датафрейм в виде красивой таблицы в консоль с использованием формата fancy_grid.

    Parameters:
    - df (pd.DataFrame): Датафрейм для вывода.

    Returns:
    - None
    """
    print(df.to_markdown(tablefmt="fancy_grid"))


def clean_date(val):
    def epoch_time(val, default_time):
        djd = datetime.timedelta(val)
        dublin_epoch = datetime.datetime(1899, 12, 30, 12)
        res = dublin_epoch + djd
        new_val = res.strftime("%Y-%m-%d")
        new_val = datetime.datetime.strptime(str(new_val), '%Y-%m-%d')
        if str(new_val) == '2000-01-01':
            new_val = default_time
        return new_val

    def string_time(val):
        val = val.replace(' ', '').replace('г.', '').replace('г', '').replace('.', '-').replace(',', '-')
        def_val = val
        new_val = False

        try:
            new_val = datetime.datetime.strptime(str(val), '%Y-%m-%d %H:%M:%S')
        except:
            try:
                new_val = datetime.datetime.strptime(str(val), '%Y-%m-%d')
            except:
                try:
                    new_val = datetime.datetime.strptime(str(val), '%d-%m-%Y')
                except:
                    try:
                        if '-' in val:
                            temp = val.split('-')
                            if len(temp[-1]) == 2:
                                year = f'20{temp[-1]}'
                            else:
                                year = temp[-1]
                            new_val = f'{year}-{temp[0]}-{temp[1]}'
                            new_val = datetime.datetime.strptime(new_val, '%d-%m-%Y')
                        else:
                            if len(val) == 8:
                                val = f'{val[4:]}-{val[:1]}-{val[2:3]}'
                                new_val = datetime.datetime.strptime(val, '%d-%m-%Y')
                    except:
                        pass
        if not new_val:
            new_val = def_val
        return new_val

    default_time = datetime.datetime.strptime('1970-01-01 00:00:00', '%Y-%m-%d %H:%M:%S')
    new_val = False

    try:
        if type(val) == pd._libs.tslibs.timestamps.Timestamp or type(val) == datetime.datetime:
            val = val.replace(microsecond=0)
            val = datetime.datetime.strptime(str(val).split('.')[0], '%Y-%m-%d %H:%M:%S')
            new_val = datetime.datetime.strptime(str(val), '%Y-%m-%d %H:%M:%S')
        elif type(val) == str:
            if val in ['0', '', ' ', '0000-00-00 00:00:00', '0000-00-00']:
                new_val = default_time
            else:
                new_val = string_time(val)
        elif type(val) == int or type(val) == float:
            if val == 0:
                new_val = default_time
            else:
                new_val = epoch_time(val, default_time)
        elif val == False or val == None or type(val) == pd._libs.tslibs.nattype.NaTType or val is np.nan:
            new_val = default_time
    except Exception as err:
        pass
        # print(val, type(val), err)
    if not new_val or type(new_val) != datetime.datetime:
        new_val = default_time
        # print(val, type(val))

    # # print(new_val)
    # if new_val == default_time or str(new_val) == '1970-01-01 00:00:00':
    #     # print(new_val)
    #     new_val = np.nan
    return new_val


@log_decorator(logger=logger)
def types_from_df_to_df(df1, df2):
    return df2.apply(lambda col: col.astype(df1[col.name].dtype) if col.name in df1.columns else col)


@log_decorator(logger=logger)
def create_connectors(credentials_path):
    init = {}

    # Прочитать данные из файла
    credentials = LocalFile.json_to_dict(credentials_path)

    # Проверить, удалось ли прочитать данные из файла
    if not credentials:
        print(f"Error: File not found or empty at {credentials_path}")
        return init

    # Обработка данных из файла
    for name, creds in credentials.items():
        # Обработать данные для каждого коннектора
        if 'type' in creds:
            if creds['type'] == 'database':
                # Для коннектора базы данных создаем экземпляр класса DataBase
                db_name = name
                db_type = creds.get('subtype')
                host = creds.get('host')
                port = creds.get('port')
                username = creds.get('login')
                password = creds.get('password')
                init[name] = DataBase(db_type, db_name, host, port, username, password)

            elif creds['type'] == 'api':
                # Для коннектора API проверяем тип аутентификации и создаем соответствующий объект
                if creds.get('auth_type') == 'token':
                    init[name] = {
                        'host': creds.get('host'),
                        'token': creds.get('password')
                    }
                elif creds.get('auth_type') == 'logopass':
                    init[name] = {
                        'host': creds.get('host'),
                        'login': creds.get('login'),
                        'password': creds.get('password')
                    }
                elif creds.get('auth_type') == 'file' and 'Google' in name:
                    init[name] = {
                        'web_path': creds.get('host'),
                        'file_id': creds.get('host').replace('https://docs.google.com/spreadsheets/d/', '').split('/')[
                            0],
                        'access_file': creds.get('password'),
                    }
    return init


@log_decorator(logger=logger)
def convert_data_types(source, target, src_type=False):
    src_type = str(src_type)
    # Пример сопоставления типов данных между различными базами данных и Pandas DataFrame
    type_mapping = pd.DataFrame({
        'pandas': {
            'TEXT': 'object',
            'INTEGER': 'int64',
            'REAL': 'float64',
            'BOOL': 'bool',
            'DATETIME': 'datetime64[ns]',
            'DATE': 'datetime64[ns]',  # Добавлен тип данных date
            'TIME': 'datetime64[ns]',  # Добавлен тип данных time
            'BLOB': 'object',  # Предположим, что blob представлен как объект
            'FLOAT': 'float64',  # Добавлен тип данных float
        },
        'sqlite': {
            'TEXT': 'TEXT',
            'INTEGER': 'INTEGER',
            'REAL': 'REAL',
            'BOOL': 'INTEGER',  # SQLite использует целые числа для представления булевых значений
            'DATETIME': 'DATETIME',
            'DATE': 'DATE',  # Добавлен тип данных date
            'TIME': 'TIME',  # Добавлен тип данных time
            'BLOB': 'BLOB',
            'FLOAT': 'REAL',  # Добавлен тип данных float
        },
        'mysql': {
            'TEXT': 'TEXT',
            'INTEGER': 'INT',
            'REAL': 'DOUBLE',
            'BOOL': 'TINYINT',  # MySQL использует TINYINT для представления булевых значений
            'DATETIME': 'DATETIME',
            'DATE': 'DATE',  # Добавлен тип данных date
            'TIME': 'TIME',  # Добавлен тип данных time
            'BLOB': 'BLOB',
            'FLOAT': 'DOUBLE',  # Добавлен тип данных float
        },
        'mssql': {
            'TEXT': 'NVARCHAR(MAX)',
            'INTEGER': 'INT',
            'REAL': 'FLOAT',
            'BOOL': 'BIT',  # MSSQL использует BIT для представления булевых значений
            'DATETIME': 'DATETIME',
            'DATE': 'DATE',  # Добавлен тип данных date
            'TIME': 'TIME',  # Добавлен тип данных time
            'BLOB': 'VARBINARY(MAX)',
            'FLOAT': 'FLOAT',  # Добавлен тип данных float
        },
        'postgresql': {
            'TEXT': 'TEXT',
            'INTEGER': 'INTEGER',
            'REAL': 'REAL',
            'BOOL': 'BOOLEAN',  # PostgreSQL имеет отдельный тип BOOLEAN для булевых значений
            'DATETIME': 'TIMESTAMP',
            'DATE': 'DATE',  # Добавлен тип данных date
            'TIME': 'TIME',  # Добавлен тип данных time
            'BLOB': 'BYTEA',
            'FLOAT': 'REAL',  # Добавлен тип данных float
        },
        'clickhouse': {
            'TEXT': 'String',
            'INTEGER': 'Int32',
            'REAL': 'Float32',
            'BOOL': 'Int8',  # ClickHouse использует Int8 для представления булевых значений
            'DATETIME': 'DateTime',
            'DATE': 'Date',  # Добавлен тип данных date
            'TIME': 'DateTime',  # Добавлен тип данных time
            'BLOB': 'String',  # Пример, замените на соответствующий тип в ClickHouse
            'FLOAT': 'Float32',  # Добавлен тип данных float
        },
    })

    # Создаем словарь сопоставления для указанных исходного и целевого источников данных
    result = type_mapping[[source, target]]

    if src_type:
        return result.loc[result[source] == src_type].to_dict('records')[0][target]
    return result


@log_decorator(logger=logger)
def compare_df_old(old_data: pd.DataFrame, new_data: pd.DataFrame, compare_cols=[], exclude_columns=[],
                   return_data=False):
    # print(old_data)
    # print(new_data)
    exclude_columns.append('add_time')
    exclude_columns.append('update_time')
    if not compare_cols:
        # Производит проверку по всем столбцам кроме системных если не указано по каким
        compare_cols = [col for col in list(new_data.columns) if col not in exclude_columns]
    all_cols = [col for col in list(old_data.columns) if col not in exclude_columns]
    # print(all_cols)
    # Выполняем слияние данных, но сверку производим только по выбранным колонкам
    merged_data = pd.merge(old_data[all_cols], new_data[all_cols], how='outer', indicator=True, on=compare_cols)
    # print(merged_data)
    merged_data['change'] = 0

    for col in old_data[all_cols].columns:
        if col not in compare_cols:
            merged_data['change'] += merged_data[f'{col}_x'] != merged_data[f'{col}_y']

    merged_data['change'] = merged_data['change'].fillna(0)
    merged_data['change'] = (merged_data['change'] > 0) | (merged_data['_merge'] != 'both')
    need_cols = compare_cols + ['_merge', 'change']
    merged_data = merged_data[need_cols]
    # print(merged_data)

    new = []
    for i, s in merged_data.iterrows():
        data = new_data
        for column, value in dict(s).items():
            # print(column,value)
            if column not in ['_merge', 'change']:
                data = data[data[column] == value]
        # print(dict(s))
        if not data.empty:
            data = data.to_dict('records')[0]
            data.update({'_merge': s['_merge'], 'change': s['change']})
            new.append(data)
    merged_data = pd.DataFrame(new)
    df_print(merged_data)

    # Разделяем данные на три датафрейма
    only_old = merged_data[merged_data['_merge'] == 'left_only'].drop('_merge', axis=1).drop('change', axis=1)
    only_new = merged_data[merged_data['_merge'] == 'right_only'].drop('_merge', axis=1).drop('change', axis=1)
    both = merged_data[(merged_data['_merge'] == 'both')].drop('_merge', axis=1).drop('change', axis=1)
    both_changed = merged_data[(merged_data['_merge'] == 'both') &
                               (merged_data['change'] == True)].drop('_merge', axis=1).drop('change', axis=1)
    both_not_changed = merged_data[(merged_data['_merge'] == 'both') &
                                   (merged_data['change'] == False)].drop('_merge', axis=1).drop('change', axis=1)
    return False
    if not return_data:
        return merged_data
    elif return_data == 'only_old':
        return only_old
    elif return_data == 'only_new':
        return only_new
    elif return_data == 'both':
        return both
    elif return_data == 'both_changed':
        return both
    elif return_data == 'both_not_changed':
        return both
    elif return_data == 'all':
        return {'only_old': only_old,
                'only_new': only_new,
                'both': both,
                'both_changed': both_changed,
                'both_not_changed': both_not_changed,
                'compare_cols': compare_cols}


@log_decorator(logger=logger)
def compare_df(old_data: pd.DataFrame, new_data: pd.DataFrame, compare_cols=[], exclude_columns=[], return_data=False):
    if not exclude_columns:
        exclude_columns.append('add_time')
        exclude_columns.append('update_time')

    if not compare_cols:
        # Производит проверку по всем столбцам кроме системных если не указано по каким
        compare_cols = [col for col in list(new_data.columns) if col not in exclude_columns]
    # all_cols = [col for col in list(new_data.columns) if col not in exclude_columns]
    # all_cols = list(new_data.columns)
    #
    # for col in old_data.columns:
    #     if col in new_data.columns:
    #         new_data[col] = pd.Series(dtype=old_data[col].dtype)

    # Выполняем слияние данных, но сверку производим только по выбранным колонкам
    merged_data = pd.merge(old_data, new_data, how='outer', indicator=True, on=compare_cols)

    # Собираем после сверки
    need_cols = compare_cols + ['_merge']
    res = []
    for i, s in merged_data.iterrows():
        temp = {}
        for k, v in dict(s).items():
            if k in need_cols:
                temp.update({k: v})
            elif k.endswith('_x') and s['_merge'] in ['both', 'left_only']:
                temp.update({k[:-2]: v})
            elif k.endswith('_y') and s['_merge'] == 'right_only':
                temp.update({k[:-2]: v})
        res.append(temp)
    merged_data = pd.DataFrame(res)

    # Разделяем данные на три датафрейма
    only_old = merged_data[merged_data['_merge'] == 'left_only'].drop('_merge', axis=1)
    only_new = merged_data[merged_data['_merge'] == 'right_only'].drop('_merge', axis=1)
    both = merged_data[(merged_data['_merge'] == 'both')].drop('_merge', axis=1)

    if not return_data:
        return merged_data
    elif return_data == 'only_old':
        return only_old
    elif return_data == 'only_new':
        return only_new
    elif return_data == 'both':
        return both
    elif return_data == 'both_changed':
        return both
    elif return_data == 'both_not_changed':
        return both
    elif return_data == 'all':
        return {'only_old': only_old[:5],
                'only_new': only_new[:5],
                'both': both[:5],
                'compare_cols': compare_cols}


@log_decorator(logger=logger)
def search_param_value(df, search_col, search_val, param_col=False):
    if param_col:
        return df[param_col][df[search_col] == search_val].values[0]
    else:
        return dict(df[df[search_col] == search_val].iloc[0])


@log_decorator(logger=logger)
def set_columns_from_row(df, n):
    """
    Назначает n-ую строку в качестве названий столбцов и удаляет эту строку с пересчетом индексов.

    Parameters:
    - df (pd.DataFrame): DataFrame.
    - n (int): Номер строки, которую следует использовать в качестве названий столбцов.

    Returns:
    - pd.DataFrame: DataFrame с новыми названиями столбцов и пересчитанными индексами.
    """
    df = df.set_axis(df.iloc[n], axis=1, copy=False)
    df = df.iloc[n + 1:]  # Выбираем строки с индексами больше n
    df.reset_index(drop=True, inplace=True)  # Пересоздаем индекс
    return df


@log_decorator(logger=logger)
def translate_dataframe(df, to_exclude=[], tr_columns=True, tr_values=True, source_lang='ru', target_lang='en',
                        get_tech_name=False):
    """
        Статический метод для перевода значений в столбцах DataFrame.

        Parameters:
        - df (pd.DataFrame): DataFrame, в котором нужно произвести перевод.
        - columns_to_exclude (list): Список столбцов, значения в которых не нужно переводить.
        - source_lang (str): Язык исходного текста (по умолчанию 'ru').
        - target_lang (str): Язык, на который нужно перевести текст (по умолчанию 'en').

        Returns:
        - pd.DataFrame: DataFrame с переведенными значениями.
    """
    replace_values = {'Да': True, 'Нет': False, 'yes': True, 'no': False, np.nan: False}
    unique_values = []

    translated_df = df.copy()
    translated_df.replace(replace_values, inplace=True)

    for column in df.columns:
        if column not in to_exclude:
            unique_values += list(df[column].unique())

    if tr_columns:
        columns_translation = {
            el: TextProcessor().translate(el, source_lang=source_lang, target_lang=target_lang,
                                          get_tech_name=get_tech_name) for el in
            list(df.columns)}

        translated_df = translated_df.rename(columns=columns_translation)
    if tr_values:
        for text in unique_values:
            tr_text = TextProcessor().translate(text, get_tech_name=True)
            if text != tr_text:
                replace_values.update({text: tr_text})
        translated_df.replace(replace_values, inplace=True)

    return translated_df


@log_decorator(logger=logger)
def start_in_process(tasks):
    if isinstance(tasks, dict):
        for task, args in tasks.items():
            process = multiprocessing.Process(target=task, args=args)
            process.daemon = True
            process.start()
    elif isinstance(tasks, list):
        for task in tasks:
            process = multiprocessing.Process(target=task)
            process.daemon = True
            process.start()
    else:
        process = multiprocessing.Process(target=tasks)
        process.daemon = True
        process.start()


""" Вспомогательные классы """


# Класс для работы с файлами, включающий методы для проверки, создания, удаления, получения расширения и размера файла.
class FileUtils:
    # @log_decorator(logger=logger)
    @staticmethod
    def file_exists(file_path):
        """
        Проверяет, существует ли файл.

        Parameters:
        - file_path (str): Путь к файлу.

        Returns:
        - bool: True, если файл существует, False в противном случае.
        """
        return os.path.exists(file_path)

    @staticmethod
    def get_files_list(dir_path, file_pattern=False):
        if os.path.exists(dir_path):
            if file_pattern:
                file_pattern = f'{dir_path}/{file_pattern}'
                return glob.glob(file_pattern)
            return os.listdir(dir_path)
        return list()

    @log_decorator(logger=logger)
    @staticmethod
    def create_file(file_path):
        """
        Создает файл.

        Parameters:
        - file_path (str): Путь к файлу.

        Returns:
        - bool: True, если файл успешно создан, False в противном случае.
        """
        try:
            with open(file_path, 'w'):
                pass
            return True
        except Exception as e:
            print(f"Ошибка при создании файла: {e}")
            return False

    @staticmethod
    def create_directory(directory_path):
        """
        Создает директорию, если она не существует.

        Parameters:
        - directory_path (str): Путь к директории.

        Returns:
        - bool: True, если директория успешно создана или уже существует, False в противном случае.
        """
        try:
            if not os.path.exists(directory_path):
                os.makedirs(directory_path)

            return True
        except Exception as e:
            print(f"Ошибка при создании директории: {e}")
            return False

    @log_decorator(logger=logger)
    @staticmethod
    def delete_file(path, is_folder=False):
        """
        Удаляет файл.

        Parameters:
        - file_path (str): Путь к файлу.

        Returns:
        - bool: True, если файл успешно удален, False в противном случае.
        """
        if not is_folder:
            if os.path.exists(path):
                os.remove(path)
        else:
            if os.path.exists(path):
                shutil.rmtree(path)

    @log_decorator(logger=logger)
    @staticmethod
    def get_file_extension(file_path):
        """
        Получает расширение файла.

        Parameters:
        - file_path (str): Путь к файлу.

        Returns:
        - str: Расширение файла.
        """
        return os.path.splitext(file_path)[1]

    @staticmethod
    def get_file_size(file_path):
        """
        Получает размер файла в байтах.

        Parameters:
        - file_path (str): Путь к файлу.

        Returns:
        - int: Размер файла в байтах.
        """
        return os.path.getsize(file_path)


# Класс для работы с локальными файлами
class LocalFile:

    @staticmethod
    @log_decorator(logger=logger)
    def dict_to_yaml(data, path):
        """
        Метод для преобразования словаря в YAML и записи данных в файл.

        Parameters:
        - data (dict): Словарь данных.
        - path (str): Путь к файлу YAML.
        """
        with open(path, 'w', encoding='utf-8') as o_file:
            yaml.dump(data, o_file, default_flow_style=False, allow_unicode=True)

    @staticmethod
    @log_decorator(logger=logger)
    def yaml_to_dict(path):
        """
        Метод для чтения данных из файла YAML и преобразования их в словарь.

        Parameters:
        - path (str): Путь к файлу YAML.

        Returns:
        - dict: Словарь данных, прочитанных из файла.
        """
        with open(path, encoding='utf-8') as o_file:
            res = dict(yaml.load(o_file, Loader=SafeLoader))
        return res

    # @staticmethod
    # @log_decorator(logger=logger)
    # def read_excel_file(file_path, sheet_names=None, data_to_dict=False):
    #     """
    #     Метод для чтения данных из файла Excel.
    #
    #     Parameters:
    #     - file_path (str): Путь к файлу Excel.
    #     - sheet_names (list): Список имен листов для чтения.
    #
    #     Returns:
    #     - dict: Словарь данных, прочитанных из файла Excel.
    #     """
    #     if sheet_names is None:
    #         data = pd.read_excel(file_path, sheet_name=None)
    #     else:
    #         data = pd.read_excel(file_path, sheet_name=sheet_names)
    #
    #     if data_to_dict:
    #         reformat_data = {}
    #         for k, v in data.items():
    #             reformat_data.update({k: v.to_dict('records')})
    #         data = reformat_data
    #     return data

    @staticmethod
    @log_decorator(logger=logger)
    def read_excel_file(file_path, sheet_names=None, data_to_dict=False):
        """
        Метод для чтения данных из файла Excel.

        Parameters:
        - file_path (str): Путь к файлу Excel.
        - sheet_names (list): Список имен листов для чтения.

        Returns:
        - dict: Словарь данных, прочитанных из файла Excel.
        """
        with pd.ExcelFile(file_path) as xls:
            if sheet_names is None:
                data = {sheet_name: xls.parse(sheet_name) for sheet_name in xls.sheet_names}
            else:
                data = {sheet_name: xls.parse(sheet_name) for sheet_name in sheet_names}

        if data_to_dict:
            reformat_data = {}
            for k, v in data.items():
                reformat_data.update({k: v.to_dict('records')})
            data = reformat_data
        return data

    @staticmethod
    @log_decorator(logger=logger)
    def write_excel_file(data_dict, file_path):
        """
        Метод для записи данных в файл Excel.

        Parameters:
        - data_dict (dict): Словарь данных для записи.
        - file_path (str): Путь к файлу Excel.
        """
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            for sheet_name, df in data_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    @staticmethod
    @log_decorator(logger=logger)
    def json_to_dict(file_path):
        """
        Метод для чтения данных из файла JSON и преобразования их в словарь.

        Parameters:
        - file_path (str): Путь к файлу JSON.

        Returns:
        - dict: Словарь данных, прочитанных из файла JSON.
        """
        with open(file_path, 'r', encoding='utf-8') as json_file:
            data_dict = json.load(json_file)
        return data_dict

    @staticmethod
    @log_decorator(logger=logger)
    def dict_to_json(data_dict, file_path):
        """
        Метод для преобразования словаря в JSON и записи данных в файл.

        Parameters:
        - data_dict (dict): Словарь данных.
        - file_path (str): Путь к файлу JSON.
        """
        with open(file_path, 'w', encoding='utf-8') as json_file:
            json.dump(data_dict, json_file, indent=2, ensure_ascii=False)

    @staticmethod
    @log_decorator(logger=logger)
    def read_binary_file(file_path):
        """
        Метод для чтения бинарных данных из файла.

        Parameters:
        - file_path (str): Путь к бинарному файлу.

        Returns:
        - bytes: Бинарные данные из файла.
        """
        with open(file_path, 'rb') as binary_file:
            content = binary_file.read()
            return content

    @staticmethod
    @log_decorator(logger=logger)
    def write_binary_file(data, file_path):
        """
        Метод для записи бинарных данных в файл.

        Parameters:
        - data (bytes): Бинарные данные.
        - file_path (str): Путь к бинарному файлу.
        """
        with open(file_path, 'wb') as binary_file:
            binary_file.write(data)


# Класс для работы с текстом, включающий методы для перевода, транслитерации и очистки текста.
class TextProcessor:
    # @log_decorator(logger=logger)
    def translate_not_work(self, text, source_lang='ru', target_lang='en', get_tech_name=False):
        """
        Статический метод для перевода текста.

        Parameters:
        - text (str): Исходный текст для перевода.
        - source_lang (str): Язык исходного текста (по умолчанию 'ru').
        - target_lang (str): Язык, на который нужно перевести текст (по умолчанию 'en').

        Returns:
        - str: Переведенный текст.
        """
        translator = Translator()
        if text in [True, False, None, np.nan] \
                or isinstance(text, (datetime.datetime, pd.Timestamp)) \
                or not self.has_russian(text):
            return text
        translation = translator.translate(text, src=source_lang, dest=target_lang).text
        if get_tech_name:
            translation = self.get_tech_name(translation)
        return translation

    @log_decorator(logger=logger)
    def translate(self, text, source_lang='ru', target_lang='en', get_tech_name=False):
        """
        Статический метод для перевода текста.

        Parameters:
        - text (str): Исходный текст для перевода.
        - source_lang (str): Язык исходного текста (по умолчанию 'ru').
        - target_lang (str): Язык, на который нужно перевести текст (по умолчанию 'en').

        Returns:
        - str: Переведенный текст.
        """
        if text in [True, False, None, np.nan] \
                or isinstance(text, (datetime.datetime, pd.Timestamp)) \
                or not self.has_russian(text):
            return text

        translator = Translator(from_lang=source_lang, to_lang=target_lang)
        translation = translator.translate(text)

        if get_tech_name:
            translation = self.get_tech_name(translation)

        return translation

    @staticmethod
    @log_decorator(logger=logger)
    def transliterate(text, target_lang='en'):
        """
        Статический метод для транслитерации текста.

        Parameters:
        - text (str): Исходный текст для транслитерации.
        - target_lang (str): Язык транслитерации (по умолчанию 'en' для английского).

        Returns:
        - str: Транслитерированный текст.
        """
        return transliterate.translit(text, target_lang, reversed=True)

    @staticmethod
    @log_decorator(logger=logger)
    def clean_text(text, symbols_to_remove=[]):
        """
        Статический метод для очистки текста от заданных символов.

        Parameters:
        - text (str): Исходный текст.
        - symbols_to_remove (list): Список символов для удаления (по умолчанию []).

        Returns:
        - str: Очищенный текст.
        """
        if symbols_to_remove:
            for symbol in symbols_to_remove:
                text = text.replace(symbol, '')
        return text

    @staticmethod
    @log_decorator(logger=logger)
    def has_russian(text):
        # Проверяет, есть ли в тексте русские буквы
        return bool(re.search('[а-яА-Я]', text))

    @log_decorator(logger=logger)
    def get_tech_name(self, text: str = None, translit=False) -> str:
        """
        Генерирует техническое имя (slug) на основе переданного текста.

        Параметры:
        - text (str): Текст на русском или английском языке.

        Возвращает:
        - str: Техническое имя (slug).

        Пример использования:
        get_tech_name("Пример текста на русском")
        'primer_teksta_na_russkom'
        get_tech_name("Example text in English")
        'example_text_in_english'
        """
        text = str(text).lower()
        replace_data = {
            '\n': '_',
            ' \n': '_',
            '\n ': '_',
            "'": '',
            "`": '',
            ' ': '_',
            '.': '',
            ",": "_",
            '"': "_",
            "ь": '',
            "ъ": '',
            "|": '',
            "/": '_',
            "(": '',
            ")": '',
            "+": '',
            "-": '',
            "=": '',
            ";": '_',
            ":": '_',
            "?": '',
            "№": 'nomer',
            "#": '_',
            "__": '_',
        }
        for el, replacement in replace_data.items():
            text = text.replace(el, replacement)
        if translit:
            return self.transliterate(text.lower())
        return text.lower()


# Класс для работы с требованиями проекта
class ProjectRequirements:
    @log_decorator(logger=logger)
    def __init__(self, project_path):
        self.project_path = project_path
        self.requirements_file = f'{project_path}/requirements.txt'

    @log_decorator(logger=logger)
    def determine_required_libraries(self, modules):
        """
        Метод для определения необходимых библиотек.

        Parameters:
        - modules (list): Список модулей, которые необходимо проверить.

        Returns:
        - list: Список необходимых библиотек.
        """
        required_libraries = []

        # Перебор модулей и попытка их импорта
        for module in modules:
            try:
                importlib.import_module(module)
            except ImportError:
                # Если модуль не может быть импортирован, добавляем его в список необходимых
                required_libraries.append(module)

        return required_libraries

    @log_decorator(logger=logger)
    def generate_requirements_file(self, required_libraries):
        """
        Метод для генерации файла requirements.txt на основе списка необходимых библиотек.

        Parameters:
        - required_libraries (list): Список необходимых библиотек.
        """
        with open(self.requirements_file, 'w') as file:
            file.write('\n'.join(required_libraries))

    @log_decorator(logger=logger)
    def install_requirements(self):
        """
        Метод для установки необходимых библиотек с использованием pip.
        """
        subprocess.call([sys.executable, '-m', 'pip', 'install', '-r', self.requirements_file])

    @log_decorator(logger=logger)
    def old_determine_required_libraries(self, modules=[]):
        required_libraries = []

        for module in modules:
            try:
                importlib.import_module(module)
            except ImportError:
                required_libraries.append(module)

        return required_libraries

    @log_decorator(logger=logger)
    def old_generate_requirements_file(self, required_libraries):
        with open(self.requirements_file, 'w') as file:
            file.write('\n'.join(required_libraries))

    @log_decorator(logger=logger)
    def old_detect_installed_libraries(self):
        installed_libraries = subprocess.check_output([sys.executable, '-m', 'pip', 'freeze'], universal_newlines=True)
        return installed_libraries.strip().split('\n')

    @log_decorator(logger=logger)
    def old_install_requirements(self):
        subprocess.call([sys.executable, '-m', 'pip', 'install', '-r', self.requirements_file])


""" Классы для API """


# Класс для работы с Google Drive API через google-api-python-client
class GoogleDriveAPI:
    '''
    Класс для работы с Google Drive API через google-api-python-client.
    Документация: http://datalytics.ru/all/rabotaem-s-api-google-drive-s-pomoschyu-python/
    '''

    def __init__(self, service_account_file: str):
        # Инициализация класса с использованием файла учетных данных
        scopes = ['https://www.googleapis.com/auth/drive']
        credentials = service_account.Credentials.from_service_account_file(service_account_file, scopes=scopes)
        self.service = build('drive', 'v3', credentials=credentials)

    @log_decorator(logger=logger)
    def get_file_id(self, ref):
        return ref.replace('https://docs.google.com/spreadsheets/d/', '').split('/')[0]

    @log_decorator(logger=logger)
    def create_folder(self, base_folder_id, folder_name):
        """Создание папки в указанной базовой папке."""
        # Метаданные новой папки
        file_metadata = {'name': folder_name,
                         'mimeType': 'application/vnd.google-apps.folder',
                         'parents': [base_folder_id]}
        # Создание запроса для создания папки
        file = self.service.files().create(body=file_metadata, fields='id').execute()
        return file.get('id')

    @log_decorator(logger=logger)
    def upload_to_folder(self, file_path, base_folder_id=None):
        """Загрузка файла в указанную папку."""
        # Создание метаданных файла
        if base_folder_id:
            file_metadata = {'name': file_path,
                             'parents': [base_folder_id]}
            media = MediaFileUpload(file_path, resumable=True)
        else:
            file_metadata = {'name': file_path}
            media = MediaFileUpload(file_path)
        # Создание запроса для загрузки файла
        file = self.service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        return file.get('id')

    @log_decorator(logger=logger)
    def move_file_to_folder(self, file_id, folder_id):
        """Перемещение файла в указанную папку."""
        # Получение существующих родительских папок
        file = self.service.files().get(fileId=file_id, fields='parents').execute()
        previous_parents = ",".join(file.get('parents'))
        # Перемещение файла в новую папку
        file = self.service.files().update(fileId=file_id, addParents=folder_id,
                                           removeParents=previous_parents,
                                           fields='id, parents').execute()
        return file.get('parents')

    @log_decorator(logger=logger)
    def download_file(self, real_file_id, file_path, google_create=True):
        """Загрузка файла."""
        if google_create:
            mime_types = {
                '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            }
        else:
            mime_types = {
                '.pdf': 'application/pdf',
                '.txt': 'text/plain'
            }

        extension = os.path.splitext(file_path)[-1].lower()
        request_mimeType = mime_types.get(extension, None)

        if request_mimeType:
            request = self.service.files().export_media(
                fileId=real_file_id, mimeType=request_mimeType
            ) if google_create else self.service.files().get_media(fileId=real_file_id, mimeType=request_mimeType)
        else:
            request = self.service.files().get_media(fileId=real_file_id)

        with open(file_path, 'wb') as file:
            downloader = MediaIoBaseDownload(file, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()

    @log_decorator(logger=logger)
    def remove_file(self, file_id):
        """Удаление файла."""
        return self.service.files().delete(fileId=file_id).execute()

    @log_decorator(logger=logger)
    def search_file(self, parent_id, file_name):
        """Поиск файла в указанной папке."""
        results = self.service.files().list(pageSize=5,
                                            fields="nextPageToken, files(id, name, mimeType, parents, createdTime)",
                                            q=f"'{parent_id}' in parents and name contains '{file_name}'").execute()
        if results['files']:
            return results['files'][0]['id']

    @log_decorator(logger=logger)
    def block(self, file_id):
        """Блокировка файла."""
        return self.service.files().update(fileId=file_id,
                                           body={"contentRestrictions": [{"readOnly": "true",
                                                                          "reason": "Contact the administrator"}]}).execute()

    @log_decorator(logger=logger)
    def unblock(self, file_id):
        """Разблокировка файла."""
        return self.service.files().update(fileId=file_id,
                                           body={"contentRestrictions": [{"readOnly": "false"}]}).execute()

    @log_decorator(logger=logger)
    def check_block(self, file_id):
        """Проверка статуса блокировки файла."""
        return self.service.files().get(fileId=file_id, fields="contentRestrictions").execute()


# Класс для работы с Google Sheets API
class GoogleSheetsAPI:
    def __init__(self, credentials_json):
        """
        Инициализация класса GoogleSheetsAPI.

        Parameters:
        - credentials_json (str): Путь к файлу JSON с учетными данными.
        """
        self.credentials_json = credentials_json

    @log_decorator(logger=logger)
    def get_file_id(self, ref):
        return ref.replace('https://docs.google.com/spreadsheets/d/', '').split('/')[0]

    @log_decorator(logger=logger)
    def authenticate(self, file_id):
        """
        Аутентификация в Google Sheets API.

        Parameters:
        - file_id (str): Идентификатор Google Sheets файла.

        Returns:
        - gspread.Spreadsheet: Объект Google Sheets Spreadsheet.
        """
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(self.credentials_json, scope)
        gc = gspread.authorize(creds)
        return gc.open_by_key(file_id)

    @log_decorator(logger=logger)
    def create_worksheet(self, file_id, title):
        """
        Создание нового листа в Google Sheets.

        Parameters:
        - file_id (str): Идентификатор Google Sheets файла.
        - title (str): Название нового листа.

        Returns:
        - gspread.models.Worksheet: Объект Google Sheets Worksheet.
        """
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.add_worksheet(title=title)
            return worksheet

    @log_decorator(logger=logger)
    def delete_worksheet(self, file_id, title):
        """
        Удаление листа из Google Sheets.

        Parameters:
        - file_id (str): Идентификатор Google Sheets файла.
        - title (str): Название листа для удаления.

        Returns:
        - bool: True, если лист успешно удален, False в противном случае.
        """
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = self.get_worksheet_by_title(spreadsheet, title)
            if worksheet:
                spreadsheet.del_worksheet(worksheet)
                return True
        return False

    def get_worksheet_by_title(self, spreadsheet, title):
        """
        Получение листа по названию.

        Parameters:
        - spreadsheet (gspread.Spreadsheet): Объект Google Sheets Spreadsheet.
        - title (str): Название листа.

        Returns:
        - gspread.models.Worksheet: Объект Google Sheets Worksheet.
        """
        try:
            worksheet = spreadsheet.worksheet(title)
            return worksheet
        except WorksheetNotFound:
            return None

    @log_decorator(logger=logger)
    def get_data_from_sheet(self, file_id, sheet_name, cell_range=None):
        """
        Получение данных из листа Google Sheets.

        Parameters:
        - file_id (str): Идентификатор Google Sheets файла.
        - sheet_name (str): Название листа.
        - cell_range (str): Диапазон ячеек (по умолчанию None).

        Returns:
        - pd.DataFrame: DataFrame с данными из Google Sheets.
        """
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.worksheet(sheet_name)
            if cell_range:
                data = pd.DataFrame.from_dict(worksheet.get(cell_range))
            else:
                data = pd.DataFrame.from_dict(worksheet.get_all_records(
                    expected_headers=[],
                    empty2zero=False,
                    head=1,
                    default_blank='',
                    allow_underscores_in_numeric_literals=True,
                    numericise_ignore=None,
                    value_render_option='UNFORMATTED_VALUE'
                ))
            return data

    @log_decorator(logger=logger)
    def upload_data_to_sheet(self, file_id, sheet_name, data, start_cell="A1"):
        """
        Загрузка данных в лист Google Sheets.

        Parameters:
        - file_id (str): Идентификатор Google Sheets файла.
        - sheet_name (str): Название листа.
        - data (str): Данные для загрузки.
        - start_cell (str): Начальная ячейка для загрузки данных (по умолчанию "A1").
        """
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.worksheet(sheet_name)
            worksheet.update(start_cell, data)

    @log_decorator(logger=logger)
    def upload_data_from_dict_to_sheet(self, file_id, sheet_name, data_dict):
        """
        Загрузка данных в лист Google Sheets.

        Parameters:
        - file_id (str): Идентификатор Google Sheets файла.
        - sheet_name (str): Название листа.
        - data_dict (dict): Словарь, где ключ — адрес ячейки, значение — данные.
        """
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.worksheet(sheet_name)
            for cell, value in data_dict.items():
                worksheet.update(cell, value)

    @log_decorator(logger=logger)
    def upload_df_data_to_sheet(self, file_id, sheet_name, data_df):
        """
        Загрузка данных из DataFrame в лист Google Sheets.

        Parameters:
        - file_id (str): Идентификатор Google Sheets файла.
        - sheet_name (str): Название листа.
        - data_df (pd.DataFrame): DataFrame с данными для загрузки.
        - start_cell (str): Начальная ячейка для загрузки данных (по умолчанию "A1").
        """
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.worksheet(sheet_name)

            # Clear existing data in the worksheet
            worksheet.clear()

            # Resize the worksheet based on the DataFrame dimensions
            worksheet.resize(rows=data_df.shape[0], cols=data_df.shape[1])

            # Write the DataFrame to the worksheet starting from the specified cell
            set_with_dataframe(worksheet, data_df, include_index=False, include_column_header=True)

    @log_decorator(logger=logger)
    def get_cell_addresses_from_dataframe(self, df, start_cell="A1"):
        """
        Получение адресов ячеек из DataFrame.

        Parameters:
        - df (pd.DataFrame): DataFrame с данными.
        - start_cell (str): Начальная ячейка (по умолчанию "A1").

        Returns:
        - list: Список относительных адресов ячеек.
        """
        wb = Workbook()
        ws = wb.active

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        cell_addresses = [cell.coordinate for row in
                          ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, min_col=ws.min_column,
                                       max_col=ws.max_column) for cell in row]
        start_row, start_col = ws[start_cell].row, ws[start_cell].column
        relative_addresses = [f"{chr(ord(cell[0]) - ord('A') + start_col)}{cell[1] - start_row + 1}" for cell in
                              cell_addresses]

        return relative_addresses

    @log_decorator(logger=logger)
    def check_cell_value(self, file_id, sheet_name, cell_address, compare_value):
        """
        Проверка значения ячейки в Google Sheets.

        Parameters:
        - file_id (str): Идентификатор Google Sheets файла.
        - sheet_name (str): Название листа.
        - cell_address (str): Адрес ячейки.
        - compare_value (str): Значение для сравнения.

        Returns:
        - bool: Результат сравнения.
        """
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.worksheet(sheet_name)
            cell_value = worksheet.acell(cell_address).value
            return cell_value == compare_value


# Класс для работы с базой данных
class DataBase:
    def __init__(self, db_type, db_name, host=False, port=False, username=False, password=False):
        """
        Инициализация объекта базы данных.

        Parameters:
        - db_type (str): Тип базы данных.
        - database (str): Название базы данных.
        - host (str): Хост базы данных.
        - port (int): Порт базы данных.
        - username (str): Имя пользователя базы данных.
        - password (str): Пароль для доступа к базе данных.
        """
        self.db_type = db_type
        self.db_name = db_name
        self.host = host
        self.port = port
        self.username = username
        self.password = password

    ''' Вспомогательные методы '''

    @log_decorator(logger=logger)
    def get_table_columns(self, keys: dict):
        """
        Получение строкового представления колонок таблицы.

        Parameters:
        - keys (dict): Словарь с ключами.

        Returns:
        - str: Строковое представление колонок таблицы.
        """
        return str(keys). \
            replace('"', ""). \
            replace("'", ""). \
            replace(":", ""). \
            replace("{", ""). \
            replace("}", ""). \
            replace('Europe/Moscow', "'Europe/Moscow'"). \
            replace('UTC', "'UTC'")

    @log_decorator(logger=logger)
    def parsing_df_col_types(self, df):
        res = {column_name: convert_data_types('pandas', 'clickhouse', src_type=column_type) for
               column_name, column_type in df.dtypes.items()}
        df = False
        return res

    @log_decorator(logger=logger)
    def __create_engine(self, without_bd=False):
        """
        Внутренний метод для создания движка SQLAlchemy.

        Parameters:
        - without_bd (bool): Флаг для определения, создавать движок с базой данных или без.

        Returns:
        - engine: Объект SQLAlchemy Engine.
        """
        if not without_bd:
            if self.db_type == 'clickhouse':
                return clickhouse_connect.get_client(database=self.db_name, host=self.host, port=self.port,
                                                     user=self.username, password=self.password)
            else:
                if self.db_type == 'sqlite':
                    db_path = os.getcwd()
                    if self.host:
                        db_path = self.host
                    conn_str = f"sqlite:///{db_path}/{self.db_name}.db"
                    return create_engine(conn_str)
                elif self.db_type == 'mysql':
                    conn_str = f"mysql+pymysql://{self.username}:{self.password}@{self.host}:{self.port}/{self.db_name}"
                    return create_engine(conn_str)
                elif self.db_type == 'mssql':
                    conn_str = f"mssql+pyodbc://{self.username}:{self.password}@{self.host}:{self.port}/{self.db_name}?driver=ODBC+Driver+17+for+SQL+Server"
                    return create_engine(conn_str)
                elif self.db_type == 'postgres':
                    conn_str = f"postgresql+psycopg2://{self.username}:{self.password}@{self.host}:{self.port}/{self.db_name}"
                    return create_engine(conn_str)
        else:
            if self.db_type == 'clickhouse':
                return clickhouse_connect.get_client(host=self.host, port=self.port,
                                                     user=self.username, password=self.password)
            else:
                if self.db_type == 'sqlite+pysqlite':
                    conn_str = f"sqlite:///{self.db_name}"
                    return create_engine(conn_str)
                elif self.db_type == 'mysql':
                    conn_str = f"mysql+pymysql://{self.username}:{self.password}@{self.host}:{self.port}"
                    return create_engine(conn_str)
                elif self.db_type == 'mssql':
                    conn_str = f"mssql+pyodbc://{self.username}:{self.password}@{self.host}:{self.port}?driver=ODBC+Driver+17+for+SQL+Server"
                    return create_engine(conn_str)
                elif self.db_type == 'postgres':
                    conn_str = f"postgresql+psycopg2://{self.username}:{self.password}@{self.host}:{self.port}"
                    return create_engine(conn_str)
                else:
                    raise ValueError(f"Unsupported database type: {self.db_type}")

    ''' Методы для управления БД '''

    @log_decorator(logger=logger)
    def check_database_exists(self):
        """
        Проверка существования базы данных.

        Returns:
        - bool: True, если база данных существует, False в противном случае.
        """
        query = f"SHOW DATABASES LIKE '{self.db_name}'"
        return self.execute_query(query, 'exist', without_bd=True)

    @log_decorator(logger=logger)
    def create_database(self):
        """
        Создание базы данных.
        """
        if not self.check_database_exists():
            self.execute_query(f"CREATE DATABASE {self.db_name}", without_bd=True)

    @log_decorator(logger=logger)
    def drop_database(self):
        """
        Удаление базы данных.
        """
        if self.check_database_exists():
            self.execute_query(f"DROP DATABASE {self.db_name}", without_bd=True)

    ''' Методы изменения БД'''

    @log_decorator(logger=logger)
    def rename_database(self, new_name):
        """
        Переименование базы данных.

        Parameters:
        - new_name (str): Новое имя базы данных.
        """
        if self.check_database_exists():
            self.execute_query(f"RENAME DATABASE {self.db_name} TO {new_name}", without_bd=True)

    ''' Методы для управления таблицами '''

    @log_decorator(logger=logger)
    def check_table_exists(self, table_name):
        """
        Проверка существования таблицы.

        Parameters:
        - table_name (str): Имя таблицы.

        Returns:
        - bool: True, если таблица существует, False в противном случае.
        """
        query = f"SHOW TABLES LIKE '{table_name}'"
        if self.db_type == 'sqlite':
            query = f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'"
        return self.execute_query(query, 'exist')

    @log_decorator(logger=logger)
    def create_table(self, table_name, table_columns, engine=False, table_partition=False, order=False):
        """
        Создание таблицы.

        Parameters:
        - table_name (str): Имя таблицы.
        - table_columns (dict): Словарь с колонками таблицы.
        - engine (str): Тип движка таблицы (по умолчанию 'MergeTree()').
        - table_partition (str): Ключ разделения таблицы (по умолчанию 'qwe').
        - order (str): Порядок сортировки таблицы (по умолчанию 'qwe').
        """
        if not self.check_table_exists(table_name):
            first_col = list(table_columns.keys())[0]
            table_columns = self.get_table_columns(table_columns)
            if self.db_type == 'clickhouse':
                if not engine:
                    engine = 'MergeTree()'
                if not order:
                    order = first_col
                # order = get_tech_name(order)

            query = f"""CREATE TABLE 
                            {table_name} (
                                {table_columns}
                             )"""
            if engine:
                query += f' engine {engine}'
            if order:
                query += f' order by {order}'
            if table_partition:
                query += f' partition by {table_partition}'

            self.execute_query(query)

    @log_decorator(logger=logger)
    def drop_table(self, table_name):
        """
        Удаление таблицы.

        Parameters:
        - table_name (str): Имя таблицы.
        """
        if self.check_table_exists(table_name):
            self.execute_query(f"DROP TABLE {table_name}")

    ''' Методы изменения таблиц '''

    @log_decorator(logger=logger)
    def rename_table(self, old_name, new_name):
        """
        Переименование таблицы.

        Parameters:
        - old_name (str): Старое имя таблицы.
        - new_name (str): Новое имя таблицы.
        """
        if self.check_table_exists(old_name):
            self.execute_query(f"""RENAME TABLE {old_name} TO {new_name}""")

    @log_decorator(logger=logger)
    def alter_column_data_type(self, table_name, column_name, new_data_type):
        """
        Изменение типа данных колонки таблицы.

        Parameters:
        - table_name (str): Имя таблицы.
        - column_name (str): Имя колонки.
        - new_data_type (str): Новый тип данных.
        """
        if self.check_table_exists(table_name):
            query = f"ALTER TABLE {table_name} ALTER COLUMN {column_name} TYPE {new_data_type}"
            self.execute_query(query)

    @log_decorator(logger=logger)
    def add_column(self, table_name, column_name, data_type):
        """
        Добавление новой колонки в таблицу.

        Parameters:
        - table_name (str): Имя таблицы.
        - column_name (str): Имя новой колонки.
        - data_type (str): Тип данных новой колонки.
        """
        if self.check_table_exists(table_name):
            query = f"ALTER TABLE {table_name} ADD COLUMN {column_name} {data_type}"
            self.execute_query(query)

    @log_decorator(logger=logger)
    def get_info_table_columns(self, table_name):
        """
        Получение списка колонок и их типов данных из таблицы.

        Parameters:
        - table_name (str): Имя таблицы.

        Returns:
        - result_dict: Словарь, где ключи - имена колонок, значения - их типы данных.
        """
        if self.check_table_exists(table_name):
            query = f"SELECT name, type FROM system.columns WHERE table = '{table_name}'"
            result = self.execute_query(query)
            result_dict = dict(zip(result['name'], result['type']))
        return result_dict

    @log_decorator(logger=logger)
    def get_query_created_table(self, table_name, new_table_name=False):
        """
        Получение запроса создавшего таблицу.

        Parameters:
        - table_name (str): Имя таблицы.

        Returns:
        - query (str): Запрос. Если new_table_name задан, то table_name будет заменен на new_table_name.
        """
        if self.check_table_exists(table_name):
            query = f"SHOW CREATE TABLE {self.db_name}.{table_name}"
            result = self.execute_query(query, 'list')[0]['statement']
            if new_table_name:
                result = result.replace(table_name, new_table_name)
            return result

    ''' Методы изменения данных '''

    @log_decorator(logger=logger)
    def update_table_from_dict(self, table_name, data: dict, filt=False, return_query=False):
        update = []
        for k, v in data.items():
            if isinstance(v, str):
                update.append(f"{k} = '{v}'")
            elif isinstance(v, (int, float)):
                update.append(f"{k} = {v}")
            elif isinstance(v, (pd.Timestamp, datetime.datetime)):
                update.append(f"{k} = '{v}'")
        update = ', '.join(update)

        query = f'''
            ALTER TABLE
                {table_name}
            UPDATE
                {update}'''

        # print(query)

        if filt:
            where = False
            where_parts = []
            if isinstance(filt, dict):
                for k, v in filt.items():
                    if isinstance(v, str):
                        where_parts.append(f"{k} = '{v}'")
                    elif isinstance(v, (int, float, np.int64)):
                        where_parts.append(f"{k} = {v}")
                    elif isinstance(v, (pd.Timestamp, datetime.datetime)):
                        where_parts.append(f"{k} = '{v}'")
            elif isinstance(filt, list):
                if k in filt:
                    if isinstance(v, str):
                        where_parts.append(f"{k} = '{v}'")
                    elif isinstance(v, (int, float)):
                        where_parts.append(f"{k} = {v}")
                    elif isinstance(v, (pd.Timestamp, datetime.datetime)):
                        where_parts.append(f"{k} = '{v}'")
            elif isinstance(filt, str):
                where = filt
            if where_parts:
                where = ' and '.join(where_parts)
            if where:
                query += f''' 
                where {where}'''
        if return_query:
            return query
        self.execute_query(query)

    @log_decorator(logger=logger)
    def update_table_from_df(self, table_name, data: pd.DataFrame, filt=False, return_query=False):
        querys = []
        for row_dict in data.to_dict('records'):
            data = row_dict
            if filt:
                if isinstance(filt, dict):
                    filt = filt
                elif isinstance(filt, str):
                    if filt in row_dict:
                        filt = {filt: row_dict[filt]}
                    else:
                        filt = filt
                elif isinstance(filt, list):
                    filt = {key: row_dict[key] for key in row_dict if key in filt}
            querys.append(self.update_table_from_dict(table_name, data, filt, return_query))
        if return_query:
            return querys

    @log_decorator(logger=logger)
    def update_table_from_table(self, src_table, update_table, update_columns, filt):
        # Генерация SET части запроса
        if isinstance(update_columns, dict):
            set_part = ', '.join([f't1.{k} = t2.{v}' for k, v in update_columns.items()])
        elif isinstance(update_columns, list):
            set_part = ', '.join([f't1.{k} = t2.{k}' for k in update_columns])
        elif isinstance(update_columns, str):
            set_part = f', t1.{update_columns} = t2.{update_columns}'

        # Генерация WHERE части запроса
        if isinstance(update_columns, dict):
            where_part = ' AND '.join([f't1.{k} = t2.{v}' for k, v in filt.items()])
        elif isinstance(update_columns, list):
            where_part = ' AND '.join([f't1.{k} = t2.{k}' for k in filt])
        elif isinstance(update_columns, str):
            where_part = f' AND t1.{update_columns} = t2.{update_columns}'

        # Формирование SQL-запроса
        query = f'''
            UPDATE {src_table} AS t1
            SET
                {set_part}
            FROM {update_table} AS t2
            WHERE
                {where_part}
        '''
        self.execute_query(query)

    @log_decorator(logger=logger)
    def delete_from_table(self, table_name, filt=False, return_query=False):
        query = f'''
            DELETE FROM
                {table_name}'''

        if filt:
            where = False
            if isinstance(filt, dict):
                where = ' AND '.join([f"{k} = '{v}'" if isinstance(v, (str, pd.Timestamp))
                                      else f"{k} = {v}" for k, v in filt.items()])
            elif isinstance(filt, str):
                where = filt
            if where:
                query += f''' 
                    WHERE 
                        {where}'''
        if return_query:
            return query
        self.execute_query(query)

    @log_decorator(logger=logger)
    def delete_from_table_by_df(self, table_name, data: pd.DataFrame, filt=False, return_query=False):
        querys = []
        for row_dict in data.to_dict('records'):
            data = row_dict
            if filt:
                if isinstance(filt, dict):
                    filt = filt
                elif isinstance(filt, str):
                    if filt in row_dict:
                        filt = {filt: row_dict[filt]}
                    else:
                        filt = filt
                elif isinstance(filt, list):
                    filt = {key: row_dict[key] for key in row_dict if key in filt}
            querys.append(self.delete_from_table(table_name, filt, return_query))
        if return_query:
            return querys

    ''' Метод для выполнения SQL-запросов с использованием SQLAlchemy '''

    @log_decorator(logger=logger)
    def execute_query(self, query, format_return=False, without_bd=False):
        # print(query)
        """
        Выполняет SQL-запрос с использованием SQLAlchemy Engine, обрабатывает результаты
        и возвращает данные в указанном формате.

        Параметры:
        - query (str): SQL-запрос для выполнения.
        - format_return (str or bool): Формат, в который нужно преобразовать результат.
          Поддерживаемые значения: 'df' (DataFrame), 'list' (список словарей),
          'exist' (проверка наличия данных), False (без дополнительного преобразования).
        - without_bd (bool): Флаг, указывающий, нужно ли использовать базу данных.
          Если True, выполняется без подключения к базе данных.

        Возвращает:
        - В зависимости от format_return:
          - Если format_return='df', возвращает pandas DataFrame.
          - Если format_return='list', возвращает список словарей.
          - Если format_return='exist', возвращает True, если данные существуют, иначе False.
          - Если format_return=False, возвращает оригинальные данные без преобразования.

        В случае ошибки выполнения SQL-запроса возвращает None.
        """
        # Создание SQLAlchemy Engine
        engine = self.__create_engine(without_bd=without_bd)

        # Проверка типа базы данных (ClickHouse или другая)
        if self.db_type == 'clickhouse':
            # Для ClickHouse, установка лимита запросов на 0 для SELECT и SHOW запросов
            if 'select' in query.lower() or 'show' in query.lower():
                engine.query_limit = 0
                data = engine.query_df(query)

                # Если тип данных - список, преобразовать в пустой DataFrame
                if type(data) == list:
                    data = pd.DataFrame()

                # Получение названий столбцов
                columns = data.columns
                engine.close()
            else:
                # Выполнение запросов
                data = engine.command(query)
                engine.close()
        else:
            Session = sessionmaker(bind=engine)
            with Session() as session:
                # Для других баз данных (не ClickHouse), использование SQLAlchemy Session
                result = session.execute(text(query))
                try:
                    data = result.fetchall()
                    columns = result.keys()
                except:
                    data = pd.DataFrame()
                    columns = []
                # Закрытие сессии
                session.close()

        try:
            engine.close()
        except:
            try:
                session.close()
            except:
                pass

        # Обработка формата возвращаемых данных
        if format_return == 'df':
            # Преобразование в DataFrame, если указан соответствующий формат
            data = pd.DataFrame(data, columns=columns)
        elif format_return == 'list':
            # Преобразование в список словарей, если указан соответствующий формат
            data = pd.DataFrame(data, columns=columns)
            data = data.to_dict('records')
        elif format_return == 'exist':
            # Преобразование в список словарей и проверка на наличие данных
            data = pd.DataFrame(data, columns=columns)
            data = data.to_dict('records')
            if data:
                if data[0]:
                    return True
            return False
        return data

    ''' Сложные составные методы'''

    @log_decorator(logger=logger)
    def drop_view(self, view_name: str):
        db_request = f"""DROP VIEW IF EXISTS {self.db_name}.{view_name}"""
        self.execute_query(db_request)

    @log_decorator(logger=logger)
    def create_simple_view(self, view_name, query):
        self.drop_view(view_name)

        db_request = f"""
                    CREATE VIEW 
                        {self.db_name}.{view_name}
                    AS ({query})"""
        self.execute_query(db_request)

    @log_decorator(logger=logger)
    def create_materialized_view(self, view_name, query, table_name, table_columns, engine=False, table_partition=False,
                                 order=False):
        self.drop_view(view_name)
        self.create_table(table_name, table_columns, engine, table_partition, order)

        db_request = f"""
            create materialized view if not exists {self.db_name}.{view_name} to {table_name} as 
                ({query})
        """
        self.execute_query(db_request)

    @log_decorator(logger=logger)
    def create_live_view(self, view_name, query):
        self.drop_view(view_name)

        db_request = f'''
            create live view if not exists {self.db_name}.{view_name} WITH REFRESH 1 AS 
                ({query})
        '''
        self.execute_query(db_request)

    @log_decorator(logger=logger)
    def create_table_from_df(self, dataframe, table_name, table_columns=False, engine=False,
                             table_partition=False, order=False, drop_temp_table=True):
        if not table_columns:
            table_columns = self.parsing_df_col_types(dataframe)
        self.create_table(table_name, table_columns, engine=engine, table_partition=table_partition, order=order)
        self.write_data_to_table(dataframe, table_name, operation='append_all', drop_temp_table=drop_temp_table)
        dataframe = False

    @log_decorator(logger=logger)
    def substitution_table(self, table_name, drop_old=True):
        temp_table_name = f'{table_name}_temp'
        old_table_name = f'{table_name}_old'

        if self.check_table_exists(old_table_name):
            self.drop_table(old_table_name)
        query = f'RENAME TABLE {table_name} TO {old_table_name}, {temp_table_name} TO {table_name}'
        self.execute_query(query)

        if drop_old:
            if self.check_table_exists(old_table_name):
                self.drop_table(old_table_name)

    @log_decorator(logger=logger)
    def write_data_to_table(self, dataframe, table_name, operation='append', drop_temp_table=True,
                            compare_cols=[], exclude_columns=[], compare_filt_col=False, remove_old=True):
        """
        Загружает данные из датафрейма в базу данных.

        Параметры:
        - dataframe: pandas DataFrame, данные для загрузки
        - table_name: str, имя таблицы в базе данных
        - operation: str, операция загрузки ('append', 'replace', 'update')

        Варианты операции:
        - 'append': Добавить все содержимое датафрейма в БД.
        - 'replace': Заменить все содержимое БД на содержимое датафрейма.
        - 'update': Заменить только то, что есть в БД, а остальное добавить.

        Возвращает True при успешной загрузке, False в противном случае.
        """
        if self.db_type == 'clickhouse':
            engine = create_engine(f'clickhouse://{self.username}:{self.password}@{self.host}:{self.port}/{self.db_name}')
        else:
            engine = self.__create_engine()

        old_data = False
        new_data = False
        new_data = dataframe

        if not new_data.empty:
            # Если таблица не была создана ранее
            if not self.check_table_exists(table_name):
                self.create_table_from_df(new_data, table_name)
            else:
                if operation == 'append_all':
                    new_data.to_sql(table_name, engine, if_exists='append', index=False)

                if 'replace_all' in operation:
                    temp_table_name = f'{table_name}_temp'
                    if operation == 'replace_all':
                        # Проверяем наличие временной таблицы и если есть, удаляем
                        if self.check_table_exists(temp_table_name):
                            self.drop_table(temp_table_name)
                        # Создаем временную таблицу и загружаем данные
                        self.execute_query(self.get_query_created_table(table_name, new_table_name=temp_table_name))
                        new_data.to_sql(temp_table_name, engine, if_exists='append', index=False)
                        # Подменяем таблицу
                        self.substitution_table(table_name, drop_old=drop_temp_table)
                    elif operation == 'replace_all_first':
                        # Проверяем наличие временной таблицы и если есть, удаляем
                        if self.check_table_exists(temp_table_name):
                            self.drop_table(temp_table_name)
                        # Создаем временную таблицу и загружаем данные
                        self.execute_query(self.get_query_created_table(table_name, new_table_name=temp_table_name))
                        new_data.to_sql(temp_table_name, engine, if_exists='append', index=False)
                    elif operation in ['replace_all_middle', 'replace_all_end']:
                        new_data.to_sql(temp_table_name, engine, if_exists='append', index=False)
                        if operation == 'replace_all_end':
                            # Подменяем таблицу
                            self.substitution_table(table_name, drop_old=drop_temp_table)

                elif 'append_diff' in operation:
                    # Загружаем данные из основной и временной таблицы
                    query = f"SELECT * FROM {table_name}"
                    old_data = self.execute_query(query)
                    if old_data.empty:
                        new_data.to_sql(table_name, engine, if_exists='append', index=False)
                    else:
                        # Загружаем данные во временную таблицу и достаем их снова
                        temp_table_name = f'{table_name}_temp'
                        self.execute_query(self.get_query_created_table(table_name, new_table_name=temp_table_name))
                        new_data.to_sql(temp_table_name, engine, if_exists='append', index=False)
                        query = f"SELECT * FROM {temp_table_name}"
                        new_data = self.execute_query(query)
                        if self.check_table_exists(temp_table_name):
                            self.drop_table(temp_table_name)

                        diff_data = compare_df(old_data, new_data, compare_cols, exclude_columns, return_data='all')
                        # Добавляем новые
                        diff_data['only_new'].to_sql(table_name, engine, if_exists='append', index=False,
                                                     method='multi')
                        if operation == 'append_diff_with_drop':
                            # Удаляем отсутствующие в новом из старого
                            if remove_old and operation == 'append_diff_with_drop':
                                for i, s in diff_data['only_old'].iterrows():
                                    # print(self.delete_from_table(table_name, filt=dict(s), return_query=True))
                                    self.execute_query(
                                        self.delete_from_table(table_name, filt=dict(s), return_query=True))
                        diff_data = False

        try:
            engine.dispose()
            logger.log_debug(f'1 - Успешная попытка закрытия соединения {self.db_type}, table_name: {table_name}')
        except Exception as err:
            logger.log_error(
                f'1 - Неуспешная попытка закрытия соединения {self.db_type}, table_name: {table_name}, err: {err}')
            try:
                engine.close()
                logger.log_debug(f'2 - Успешная попытка закрытия соединения {self.db_type}, table_name: {table_name}')
            except Exception as err:
                logger.log_error(
                    f'2 - Неуспешная попытка закрытия соединения {self.db_type}, table_name: {table_name}, err: {err}')
        old_data = False
        new_data = False


# Класс для создания бота телеграм
class TelegramBot:
    @log_decorator(logger=logger, write_stage='TelegramBot - Init')
    def __init__(self, telegram_token=False, telegram_config_path=False, openai_token=False):
        self.telegram_token = telegram_token
        self.openai_token = openai_token
        self.config_path = telegram_config_path

        self.bot = Bot(self.telegram_token)
        self.timeout = 900
        self.updater = False
        self.dispatcher = False
        self.job_queue = False

    ''' Вспомогательные методы '''

    @log_decorator(logger=logger, write_stage='TelegramBot - get_dialog_info')
    def get_dialog_info(self, update: Update, context: CallbackContext):
        try:
            user = update.callback_query.from_user
            answer = update.callback_query.data
            chat_id = update.callback_query.message.chat_id
            msg_id = update.callback_query.message.message_id
        except:
            user = update.message.from_user
            answer = update.message.text
            chat_id = update.message.chat_id
            msg_id = update.message.message_id
        return dict(user=user, answer=answer, chat_id=chat_id, msg_id=msg_id)

    @log_decorator(logger=logger, write_stage='TelegramBot - get_config')
    def get_config(self, config_name):
        return LocalFile.json_to_dict(f'{self.config_path}/{config_name}.json')

    @staticmethod
    @log_decorator(logger=logger, write_stage='TelegramBot - update_text')
    def update_text(menu_config, dialog_info, access_user):
        text = f"{menu_config['badge']} {menu_config['text_menu']}"

        if access_user:
            user = dialog_info['user']
            replace_data = {
                '<<user_is_bot>>': str(user.is_bot),
                '<<user_id>>': str(user.id),
                '<<user_login>>': user.username,
                '<<user_first_name>>': user.first_name,
                '<<user_last_name>>': user.last_name,
                '<<user_full_name>>': access_user['full_name'],
            }
            for key, value in replace_data.items():
                if value not in [np.nan, None, False, 'Null']:
                    text = text.replace(key, value)
        return text

    @log_decorator(logger=logger, write_stage='TelegramBot - get_menu_settings')
    def get_menu_settings(self, command, update, context, callback=False):
        if callback:
            login = update.callback_query.from_user.username
        else:
            login = update.message.from_user.username
        access = self.get_config('access')[login]
        menu = self.get_config(command)
        access_menu = {'menu': menu['menu'], 'action': []}
        for elem in menu['action']:
            acc_level = access["english_access"]
            if elem[acc_level]:
                access_menu['action'].append(elem)
        return access_menu

    @log_decorator(logger=logger, write_stage='TelegramBot - create_keyboard')
    def create_keyboard(self, menu_config):
        keyboard = []
        if menu_config:
            max_row = max(item['line'] for item in menu_config) + 1
            for row in range(max_row):
                buttons_in_row = [item for item in menu_config if item['line'] == row]
                button_row = []
                for el in buttons_in_row:
                    text = f"{el['badge']} " + el['description']
                    callback = el['action']
                    button_row.append(InlineKeyboardButton(text, callback_data=callback))
                keyboard.append(button_row)
        else:
            keyboard.append(InlineKeyboardButton('🚪 Завершить', callback_data='end'))
        return InlineKeyboardMarkup(keyboard)

    @log_decorator(logger=logger, write_stage='TelegramBot - send_msg')
    def send_msg(self, text, update: Update, context: CallbackContext):
        chat_id = self.get_dialog_info(update, context)['chat_id']
        return context.bot.send_message(chat_id, text)

    @log_decorator(logger=logger, write_stage='TelegramBot - send_msg_to_channel')
    def send_msg_to_channel(self, id_channel, text, update: Update, context: CallbackContext):
        return context.bot.send_message(id_channel, text)

    @log_decorator(logger=logger, write_stage='TelegramBot - edit_msg')
    def edit_msg(self, text, keyboard, update: Update, context: CallbackContext):
        try:
            if keyboard:
                return update.message.reply_text(text, reply_markup=keyboard)
            else:
                return update.message.reply_text(text)
        except:
            if keyboard:
                return update.callback_query.edit_message_text(text, reply_markup=keyboard)
            else:
                return update.callback_query.edit_message_text(text)

    ''' Основные методы '''

    @log_decorator(logger=logger, write_stage='TelegramBot - worker')
    def worker(self, update: Update, context: CallbackContext):
        msg_info = self.get_dialog_info(update, context)
        context.user_data['prev_msg'] = msg_info
        answer = msg_info['answer']

        # При получении нового сообщения отключаем текущий таймер (если есть)
        if 'timeout_timer' in context.user_data and context.user_data['timeout_timer'] is not None:
            context.user_data['timeout_timer'].cancel()

        if answer == 'end':
            self.end(update, context)
            return 'end'
        else:
            text, keyboard, dialog_info = self.generate_menu(update, context)
            self.edit_msg(text, keyboard, update, context)

            # Запускаем новый таймер таймаута
            timeout_timer = threading.Timer(self.timeout, self.timeout_handler, args=[update, context])
            context.user_data['timeout_timer'] = timeout_timer
            timeout_timer.start()

            return dialog_info['answer']

    @log_decorator(logger=logger, write_stage='TelegramBot - generate_menu')
    def generate_menu(self, update: Update, context: CallbackContext, state=False):
        dialog_info = self.get_dialog_info(update, context)

        text = f"Добрый день! " \
               f"\n\n🤖Я бот. Но на текущий момент я не могу с вами общаться. " \
               f"Возможно у Вас нет доступа к моим функциям, либо есть технические проблемы. " \
               f"Пожалуйста обратитесь в продуктовый блок за помощью в решении Вашей проблемы. " \
               f"\n\n🙈 Спасибо за понимание!"
        keyboard = []

        if FileUtils.file_exists(f'{self.config_path}') and FileUtils.file_exists(f'{self.config_path}/start.json'):
            # Получаем настройки пользователя и меню
            if not state:
                state = dialog_info['answer']

            access_user = False
            if dialog_info['user']['username'] in self.get_config('access'):
                access_user = self.get_config('access')[dialog_info['user']['username']]
            config_menu = self.get_config(state)

            # Определяем что из меню доступно пользователю
            if access_user:
                access_menu = {'menu': config_menu['menu'], 'action': []}
                for elem in config_menu['action']:
                    acc_level = access_user["english_access"]
                    if elem[acc_level] and elem['active']:
                        access_menu['action'].append(elem)
            else:
                config_menu = self.get_config('start_without_access')
                access_menu = {'menu': config_menu['menu'], 'action': []}
                for elem in config_menu['action']:
                    if elem['active']:
                        access_menu['action'].append(elem)

            text = self.update_text(access_menu['menu'], dialog_info, access_user)
            keyboard = self.create_keyboard(access_menu['action'])
        return text, keyboard, dialog_info

    @log_decorator(logger=logger, write_stage='TelegramBot - ask_for_openai')
    def ask_for_openai(self, update: Update, context: CallbackContext):
        info = self.get_dialog_info(update, context)
        text = info['answer']

        prev_msg = context.user_data['prev_msg']
        self.updater.bot.delete_message(chat_id=prev_msg['chat_id'], message_id=prev_msg['msg_id'])

        msg = self.updater.bot.send_message(info['chat_id'], f'В OpenAI отправлен запрос. Ожидайте ответа...')
        response = OpenAIDialogManager(self.openai_token).generate_response(text)

        self.updater.bot.delete_message(chat_id=msg['chat_id'], message_id=msg['message_id'])
        msg = self.updater.bot.send_message(info['chat_id'], f'От OpenAI получен ответ: {response}')

        text, keyboard, dialog_info = self.generate_menu(update, context, state='ask_for_openai')
        msg = self.edit_msg(response, keyboard, update, context)
        context.user_data['prev_msg'] = msg

        # # Запускаем таймер таймаута
        # context.user_data['timeout_timer'] = threading.Timer(self.timeout, self.timeout_handler, args=[update, context]).start()

    @log_decorator(logger=logger, write_stage='TelegramBot - ask_for_openai_with_context')
    def ask_for_openai_with_context(self, update: Update, context: CallbackContext):
        info = self.get_dialog_info(update, context)
        text = info['answer']
        context_ztk = str(LocalFile.json_to_dict(f'{self.config_path}/context.json'))

        prev_msg = context.user_data['prev_msg']
        self.updater.bot.delete_message(chat_id=prev_msg['chat_id'], message_id=prev_msg['msg_id'])

        text = f'Используй контекст {context_ztk} и ответь: {text}'
        msg = self.updater.bot.send_message(info['chat_id'], f'В OpenAI отправлен запрос c использованием контекста. '
                                                             f'Ожидайте ответа...')
        response = OpenAIDialogManager(self.openai_token).generate_response(text)

        self.updater.bot.delete_message(chat_id=msg['chat_id'], message_id=msg['message_id'])
        text, keyboard, dialog_info = self.generate_menu(update, context, state='ask_for_openai_with_context')
        msg = self.edit_msg(response, keyboard, update, context)
        context.user_data['prev_msg'] = msg
        # Запускаем таймер таймаута
        context.user_data['timeout_timer'] = context.user_data['timeout_timer'].cancel()
        # context.user_data['timeout_timer'] = threading.Timer(self.timeout, self.timeout_handler, args=[update, context]).start()

    @log_decorator(logger=logger, write_stage='TelegramBot - end')
    def end(self, update: Update, context: CallbackContext):
        curr_msg = self.get_dialog_info(update, context)
        for i in range(2):
            try:
                self.updater.bot.delete_message(chat_id=curr_msg['chat_id'], message_id=curr_msg['msg_id'] + i)
            except:
                pass
        self.updater.bot.send_message(curr_msg['chat_id'], 'Диалог завершен по вашей инициативе')

    @log_decorator(logger=logger, write_stage='TelegramBot - timeout_handler')
    def timeout_handler(self, update, context):
        curr_msg = self.get_dialog_info(update, context)
        for i in range(2):
            try:
                self.updater.bot.delete_message(chat_id=curr_msg['chat_id'], message_id=curr_msg['msg_id'] + i)
            except:
                pass
        self.updater.bot.send_message(curr_msg['chat_id'], "Диалог завершен из-за отсутствия активности.")
        return ConversationHandler.END

    @log_decorator(logger=logger, write_stage='TelegramBot - Stop')
    def stop(self):
        try:
            self.bot.send_message(chat_id=-1002091784892, text='Завершение работы телеграм бот аналитики')
            self.updater.stop()
            self.updater.is_idle = False
        except:
            logger.log_info('Не удалось завершить работу TelegramBot')

    def run(self):
        self.updater = Updater(self.telegram_token, use_context=True)
        self.dispatcher = self.updater.dispatcher
        self.job_queue = JobQueue()
        try:
            conv_handler = ConversationHandler(
                entry_points=[CommandHandler('start', self.worker)],
                states={
                    'ask_for_openai': [
                        MessageHandler(Filters.text & ~Filters.command, self.ask_for_openai),
                    ],
                    'ask_for_openai_with_context': [
                        MessageHandler(Filters.text & ~Filters.command, self.ask_for_openai_with_context),
                    ],
                    'timeout': [
                        MessageHandler(Filters.all, self.timeout_handler),
                    ],
                },
                fallbacks=[CommandHandler('start', self.worker), CallbackQueryHandler(self.worker)],
            )

            self.dispatcher.add_handler(conv_handler)
            self.updater.job_queue = self.job_queue
            threading.Thread(target=self.updater.start_polling).start()
            self.bot.send_message(chat_id=-1002091784892, text='Запущен телеграм бот аналитики')
        except:
            self.stop()


# Класс для работы с ChatGPT
class OpenAIDialogManager:
    @log_decorator(logger=logger)
    def __init__(self, api_key):
        self.memory = []
        self.client = OpenAI(api_key=api_key)

    @log_decorator(logger=logger)
    def generate_response(self, prompt):
        try:
            # Получаем ответ от OpenAI GPT-3
            response = self.client.completions.create(
                model='gpt-3.5-turbo-instruct',
                prompt=prompt,
                temperature=0.7,
                max_tokens=3500,
            )

            # Парсим ответ
            response_dict = json.loads(response.model_dump_json(indent=5))
            response_text = response_dict['choices'][0]['text']

            # Сохраняем ответ в память
            self.memory.append(response_text)
            return response_text
        except:
            return 'Ошибка! возможно проблемы с VPN'

    @log_decorator(logger=logger)
    def get_memory(self):
        return self.memory
