import pandas as pd
import os
import gspread
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from gspread_dataframe import set_with_dataframe
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from google.oauth2 import service_account
from oauth2client.service_account import ServiceAccountCredentials


class GoogleDriveAPI:
    def __init__(self, service_account_file: str):
        scopes = ['https://www.googleapis.com/auth/drive']
        credentials = service_account.Credentials.from_service_account_file(service_account_file, scopes=scopes)
        self.service = build('drive', 'v3', credentials=credentials)

    def get_file_id(self, ref):
        return ref.replace('https://docs.google.com/spreadsheets/d/', '').split('/')[0]

    def create_folder(self, base_folder_id, folder_name):
        # Метаданные новой папки
        file_metadata = {'name': folder_name,
                         'mimeType': 'application/vnd.google-apps.folder',
                         'parents': [base_folder_id]}
        # Создание запроса для создания папки
        file = self.service.files().create(body=file_metadata, fields='id').execute()
        return file.get('id')

    def upload_to_folder(self, file_path, base_folder_id=None):
        # Создание метаданных файла
        if base_folder_id:
            file_metadata = {'name': file_path,
                             'parents': [base_folder_id]}
            media = MediaFileUpload(file_path, resumable=True)
        else:
            file_metadata = {'name': file_path}
            media = MediaFileUpload(file_path)
        # Создание запроса для загрузки файла
        file = self.service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        return file.get('id')

    def move_file_to_folder(self, file_id, folder_id):
        # Получение существующих родительских папок
        file = self.service.files().get(fileId=file_id, fields='parents').execute()
        previous_parents = ",".join(file.get('parents'))
        # Перемещение файла в новую папку
        file = self.service.files().update(fileId=file_id, addParents=folder_id,
                                           removeParents=previous_parents,
                                           fields='id, parents').execute()
        return file.get('parents')

    def download_file(self, real_file_id, file_path, google_create=True):
        if google_create:
            mime_types = {
                '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            }
        else:
            mime_types = {
                '.pdf': 'application/pdf',
                '.txt': 'text/plain'
            }

        extension = os.path.splitext(file_path)[-1].lower()
        request_mimeType = mime_types.get(extension, None)

        if request_mimeType:
            request = self.service.files().export_media(
                fileId=real_file_id, mimeType=request_mimeType
            ) if google_create else self.service.files().get_media(fileId=real_file_id, mimeType=request_mimeType)
        else:
            request = self.service.files().get_media(fileId=real_file_id)

        with open(file_path, 'wb') as file:
            downloader = MediaIoBaseDownload(file, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()

    def remove_file(self, file_id):
        return self.service.files().delete(fileId=file_id).execute()

    def search_file(self, parent_id, file_name):
        results = self.service.files().list(pageSize=5,
                                            fields="nextPageToken, files(id, name, mimeType, parents, createdTime)",
                                            q=f"'{parent_id}' in parents and name contains '{file_name}'").execute()
        if results['files']:
            return results['files'][0]['id']


class GoogleSheetsAPI:
    def __init__(self, credentials_json):
        self.credentials_json = credentials_json

    def get_file_id(self, ref):
        return ref.replace('https://docs.google.com/spreadsheets/d/', '').split('/')[0]

    def authenticate(self, file_id):
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(self.credentials_json, scope)
        gc = gspread.authorize(creds)
        return gc.open_by_key(file_id)

    def create_worksheet(self, file_id, title):
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.add_worksheet(title=title)
            return worksheet

    def delete_worksheet(self, file_id, title):
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = self.get_worksheet_by_title(spreadsheet, title)
            if worksheet:
                spreadsheet.del_worksheet(worksheet)
                return True
        return False

    def get_worksheet_by_title(self, spreadsheet, title):
        worksheet = spreadsheet.worksheet(title)
        return worksheet

    def get_data_from_sheet(self, file_id, sheet_name, cell_range=None):
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.worksheet(sheet_name)
            if cell_range:
                data = pd.DataFrame.from_dict(worksheet.get(cell_range))
            else:
                data = pd.DataFrame.from_dict(worksheet.get_all_records(
                    expected_headers=[],
                    empty2zero=False,
                    head=1,
                    default_blank='',
                    allow_underscores_in_numeric_literals=True,
                    numericise_ignore=None,
                    value_render_option='UNFORMATTED_VALUE'
                ))
            return data

    def upload_data_to_sheet(self, file_id, sheet_name, data, start_cell="A1"):
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.worksheet(sheet_name)
            worksheet.update(start_cell, data)

    def upload_data_from_dict_to_sheet(self, file_id, sheet_name, data_dict):
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.worksheet(sheet_name)
            for cell, value in data_dict.items():
                worksheet.update(cell, value)

    def upload_df_data_to_sheet(self, file_id, sheet_name, data_df):
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.worksheet(sheet_name)

            # Clear existing data in the worksheet
            worksheet.clear()

            # Resize the worksheet based on the DataFrame dimensions
            worksheet.resize(rows=data_df.shape[0], cols=data_df.shape[1])

            # Write the DataFrame to the worksheet starting from the specified cell
            set_with_dataframe(worksheet, data_df, include_index=False, include_column_header=True)

    def get_cell_addresses_from_dataframe(self, df, start_cell="A1"):
        wb = Workbook()
        ws = wb.active

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        cell_addresses = [cell.coordinate for row in
                          ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, min_col=ws.min_column,
                                       max_col=ws.max_column) for cell in row]
        start_row, start_col = ws[start_cell].row, ws[start_cell].column
        relative_addresses = [f"{chr(ord(cell[0]) - ord('A') + start_col)}{cell[1] - start_row + 1}" for cell in
                              cell_addresses]

        return relative_addresses

    def check_cell_value(self, file_id, sheet_name, cell_address, compare_value):
        spreadsheet = self.authenticate(file_id)
        if spreadsheet:
            worksheet = spreadsheet.worksheet(sheet_name)
            cell_value = worksheet.acell(cell_address).value
            return cell_value == compare_value
